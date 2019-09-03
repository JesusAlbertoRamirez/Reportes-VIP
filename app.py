from __future__ import print_function
from flask import Flask, render_template
from googleapiclient.discovery import build
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from apiclient.http import MediaIoBaseDownload
from googleapiclient import discovery
from openpyxl import load_workbook, Workbook
import pickle
import os.path
import io
import json

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/autenticar')
def autenticar():    

    creds = None
    message = ""
    # The file token.pickle stores the user's access and refresh tokens, and is
    # created automatically when the authorization flow completes for the first
    # time.
    if os.path.exists('token.pickle'):
        with open('token.pickle', 'rb') as token:
            creds = pickle.load(token)
            message = 'exists'
    # If there are no (valid) credentials available, let the user log in.
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
            message = 'refresh'
        else:
            flow = InstalledAppFlow.from_client_secrets_file(
                'credentials.json', SCOPES)            
            creds = flow.run_local_server(host='localhost',
                                        port=8080,
                                        authorization_prompt_message='Please visit this URL: {url}',
                                        success_message='The auth flow is complete; you may close this window.',
                                        open_browser=True)
            message = 'create'
            
        # Save the credentials for the next run
        with open('token.pickle', 'wb') as token:
            pickle.dump(creds, token)

    return {'msg' : message}

# If modifying these scopes, delete the file token.pickle.
SCOPES = ['https://www.googleapis.com/auth/drive']

def listFiles(creds):
    print("listFiles()")
    service = build('drive', 'v3', credentials=creds)

    # Call the Drive v3 API
    results = service.files().list(q="mimeType='application/vnd.google-apps.spreadsheet' and fullText contains '\"Matriz productos \"'",
        spaces='drive',
        pageSize=20, 
        fields="nextPageToken, files(id, name, mimeType)").execute()
    files = results.get('files', [])
    downloadFile(creds, files)   

def downloadFile(creds, files):

    service = build('drive', 'v3', credentials=creds)

    if not files:
        print('No files found.')
    else:
        print('Files:')
        for file in files:
            print(u'{0} ({1})'.format(file['name'], file['id']), file['mimeType'])

            request = service.files().export_media(fileId=file['id'],
                                             mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            fh = io.FileIO('tmp/' + file['name'] + '.xlsx', 'wb')
            downloader = MediaIoBaseDownload(fh, request)
            done = False
            while done is False:
                status, done = downloader.next_chunk()
                print ("Download %d%%." % int(status.progress() * 100))   

@app.route('/filtrar')
def search(sheet, query):

    if query=="Nombre del integrante":

        cells = sheet["A5:G23"]

        for c1,c2,c3,c4,c5,c6,c7 in cells:
            
            print("{0:8} {1:8} {2:8} {3:8} {4:8} {5:8} {6:8}". format(c1.value, c2.value, c3.value, c4.value, c5.value, c6.value, c7.value))

    elif query=="Tipo de vinculación":

        for c0,c1,c2,c3,c4,c5,c6 in cells:
            print("{1:8} {2:8}". format(c0.value, c2.value))


@app.route('/reportes/', methods=['GET'])
def manageFile():
    #book = Workbook()
    #book.save("sample.xlsx")    
    wb = load_workbook('./tmp/Matriz productos GIM.xlsx.xlsx')
    #print(wb.get_sheet_names())

    sheet = wb.active    

    query = "Nombre del integrante"
    #query = "Tipo de vinculación"

    #search(sheet, query)

    return render_template('grupos.html')

@app.route('/reportes/<string:grupo>/', methods=['GET','POST'])
def manageFileGrupo(grupo):

    print(grupo)
    wb = load_workbook('./tmp/Matriz productos '+ grupo + '.xlsx.xlsx')
    sheets = wb.get_sheet_names()

    return json.dumps(sheets)

@app.route('/reportes/<int:hoja>/', methods=['GET','POST'])
def manageFileLibro(hoja):
    print(hoja)
    wb = load_workbook('./tmp/Matriz productos GIM.xlsx.xlsx')
    sheets = wb.get_sheet_names()

    print(sheets[hoja])    

    wb.active = hoja    

    query = "Nombre del integrante"
    lista = []
    for row in wb.active.rows:
        for cell in row:
            if cell.value!=None:
                lista.append(cell.value)
                print(cell.value)

    lista.pop(0)
    return json.dumps(lista)
    
def main():    

    #listFiles(creds)     
    manageFile()

if __name__ == '__main__':
    app.run()